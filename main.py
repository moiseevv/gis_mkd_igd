import os
import openpyxl
from openpyxl import load_workbook as lwb
from datetime import datetime as dt
def create_ls(ls,i):
    new_ls = str(ls)+str(i)
    col_simv = len(new_ls)
    if col_simv == 8:
        itog_ls = "704"+str(ls)+str(i)
    if col_simv == 9:
        itog_ls = "70"+str(ls)+str(i)
    if col_simv == 10:
        itog_ls = "7"+str(ls)+str(i)
    if col_simv == 11:
        itog_ls = str(ls)+str(i)
    
    return int(itog_ls)
def replace_xl(file,dir):
    new_path = dir+"/"+file
    os.replace(file,new_path)
def create_dir():
    data_today = dt.today()
    dir_save = "a" + str(data_today).replace("-", "").replace(":", "").replace(".", "").replace(" ", "")
    os.mkdir(dir_save)
    return  dir_save
def print_hi(name):
    # Use a breakpoint in the code line below to debug your script.
    print(f'Hi, {name}')  # Press Ctrl+F8 to toggle the breakpoint.
def create_file_ls(file_item, dir_save):
    shablon_mkd = lwb('.\sample\ls_igd_06062022.xlsx')

    # чтение
    print("Читаем файл : ", file_item)
    igd = lwb(file_item)
    data_today = dt.today()
    ls = int(f"{data_today.strftime('%d')}{data_today.strftime('%m')}{data_today.strftime('%M')}{data_today.strftime('%S')}")

    if 'ИЖД' in file_item:
        tip = "ЖД"
        name_f = f".\\{dir_save}\ls_igd_{data_today.strftime('%d')}{data_today.strftime('%m')}{data_today.strftime('%Y')}.xlsx"
        ws_igd = igd["Договоры с ЖД"]
    if 'МКД' in file_item:
        tip = "МКД"
        name_f = f".\\{dir_save}\ls_mkd_{data_today.strftime('%d')}{data_today.strftime('%m')}{data_today.strftime('%Y')}.xlsx"
        ws_igd = igd["Договоры с МКД"]

    #Заполняем помещения
    ws_pomesh = shablon_mkd['Помещения']
    ws_osnovanie = shablon_mkd['Основания']
    ws_base = shablon_mkd['Основные сведения']

    for i in range(5, ws_igd.max_row + 1):
        zn_inditifikator = ws_igd.cell(i, 3).value
        zn_address = ws_igd.cell(i, 4).value

        ws_pomesh.cell(i-2,3).value = zn_inditifikator
        if 'МКД' in file_item:
            ws_pomesh.cell(i-2,4).value = "Жилое помещение"
            ws_pomesh.cell(i-2,5).value = 1
        ws_pomesh.cell(i-2,2).value = zn_address
        ws_pomesh.cell(i-2,1).value = i

        ws_base.cell(i-2,1).value = i
        ws_base.cell(i-2,2).value = create_ls(ls,i)
        ws_base.cell(i-2,4).value = "ЛС ТКО"
        ws_base.cell(i-2,21).value = 0

        ws_osnovanie.cell(i-2,1).value = i
        ws_osnovanie.cell(i-2,2).value = "Договор по обращению с ТКО (ЛС ТКО или ЛС РЦ)"
        ws_osnovanie.cell(i-2,3).value = "b38f0787-c2a2-4a1d-b646-6efb7a85580b"
        ws_osnovanie.cell(i-2,10).value = "767010"+str(i-2)
        ws_osnovanie.cell(i-2,11).value = "01.01.2019"
        ws_osnovanie.cell(i-2,12).value = "25.07.2028"

    shablon_mkd.save(name_f)
    return name_f

def create_file_ojf(file_item,dir_save):
    # шаблон
    shablon_igd = lwb('.\sample\ojf_igd_07062022.xlsx')
    ws = shablon_igd["Объекты жилищного фонда"]
    # чтение
    print("Читаем файл : ", file_item)
    igd = lwb(file_item)
    # Проверка на тип
    data_today = dt.today()

    if 'ИЖД' in file_item:
        tip = "ЖД"
        name_f = f".\\{dir_save}\ojf_igd_{data_today.strftime('%d')}{data_today.strftime('%m')}{data_today.strftime('%Y')}.xlsx"
        ws_igd = igd["Договоры с ЖД"]
    if 'МКД' in file_item:
        tip = "МКД"
        name_f = f".\\{dir_save}\ojf_mkd_{data_today.strftime('%d')}{data_today.strftime('%m')}{data_today.strftime('%Y')}.xlsx"
        ws_igd = igd["Договоры с МКД"]


    df = str(ws_igd.max_row)
    print(df)
    for i in range(5,ws_igd.max_row+1):
        zn_inditifikator = ws_igd.cell(i,3).value
        zn_address = ws_igd.cell(i,4).value

        ws.cell(i-3, 4).value = zn_inditifikator
        ws.cell(i-3, 3).value = zn_address
        ws.cell(i-3, 1).value = 'b38f0787-c2a2-4a1d-b646-6efb7a85580b'
        ws.cell(i-3, 2).value = tip
        ws.cell(i-3, 7).value = 'Размещен'

    #Сохранение ОЖФ

    shablon_igd.save(name_f)
    return name_f

# Press the green button in the gutter to run the script.



def for_igd(file_item,dir):
    name_file = create_file_ojf(file_item,dir)
    name_file_ls = create_file_ls(file_item,dir)
    print(" Создан файл - ", name_file)
    print(" Создан файл - ", name_file_ls)
    replace_xl(file_item,dir)

def check_xl(dir):
    list_files = os.listdir()
    for file_item in list_files:
        if (".xlsx" in file_item) and ('~' not in file_item):
            if ('ИЖД' in file_item) or ('igd' in file_item) or ('mkd' in file_item) or ('МКД' in file_item):
                for_igd(file_item,dir)


if __name__ == '__main__':
    print_hi('Создаем файлы загрузки')
    dir = create_dir()
    check_xl(dir)
    

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
